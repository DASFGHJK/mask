import requests
from bs4 import BeautifulSoup
import json
import pandas as pd
import datetime
import time
import random
from openpyxl import Workbook
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

today = datetime.date.today()

class WeiboTool:
    def __init__(self, cookies):
        self.weibo_content_headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'accept-encoding': 'gzip, deflate, br, zstd',
            'accept-language': 'zh-CN,zh;q=0.9',
            'cache-control': 'max-age=0',
            'cookie': cookies,
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 SLBrowser/9.0.5.12181 SLBChan/123 SLBVPV/64-bit'
        }
        self.async_headers = {
            'accept': 'application/json, text/plain, */*',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'zh-CN,zh;q=0.9',
            'client-version': 'v2.40.12',
            'cookie': cookies,
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36'
        }

    def extract_uid(self, url):
        return str(url).split('/')[-1].split('?')[0]

    def clean_text(self, text):
        return str(text).replace('\n', '').replace(' ', '').replace('\u200b', '').replace('\ue627', '')

    def get_time_ranges(self, from_time, to_time, frequency):
        from_time, to_time = pd.to_datetime(from_time), pd.to_datetime(to_time)
        time_range = list(pd.date_range(from_time, to_time, freq=f'{frequency}h'))  # 按小时分割
        if to_time not in time_range:
            time_range.append(to_time)
        return [
            [
                item.strftime('%Y-%m-%d-%H').replace('-0', '-', 1),  # 处理小时前导零
                time_range[i + 1].strftime('%Y-%m-%d-%H').replace('-0', '-', 1)
            ]
            for i, item in enumerate(time_range[:-1])
        ]

    def calculate_age(self, description):
        import re
        patterns = [
            re.compile(r'(\d{4})-(\d{1,2})-(\d{1,2})'),
            re.compile(r'(\d{4})/(\d{1,2})/(\d{1,2})'),
            re.compile(r'(\d{4})年(\d{1,2})月(\d{1,2})日'),
            re.compile(r'(\d{4})年出生'),
            re.compile(r'生日(\d{4})年(\d{1,2})月(\d{1,2})日'),
            re.compile(r'(\d{4})年(\d{1,2})月生人'),
            re.compile(r'(\d{2})岁')
        ]
        for pattern in patterns:
            match = pattern.search(description)
            if match:
                if pattern.pattern == r'(\d{4})年出生':
                    birth_date = datetime.date(int(match.group(1)), 1, 1)
                elif pattern.pattern == r'(\d{2})岁':
                    return int(match.group(1))
                else:
                    birth_date = datetime.date(*map(int, match.groups()[:3]))
                age = today.year - birth_date.year
                if today.month < birth_date.month or (today.month == birth_date.month and today.day < birth_date.day):
                    age -= 1
                return age
        return ''

    def save_to_excel(self, filename, content):
        try:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filename)
                ws = wb.active
            except FileNotFoundError:
                wb = Workbook()
                ws = wb.active
            ws.append(content)
            wb.save(filename)
            print(f'正在写入: {content}')
        except Exception as e:
            print(f"写入 Excel 文件出错: {e}")


class WeiboSpider:
    def __init__(self, url, tool, driver, max_pages=10):
        self.url = url
        self.tool = tool
        self.driver = driver
        self.max_pages = max_pages
        self.current_page = 0

    def get_content(self):
        try:
            self.driver.get(self.url)
            # 等待页面加载完成，直到找到某个特定元素
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "card-wrap"))
            )
            page_source = self.driver.page_source
            return None if '未找到' in page_source else page_source
        except Exception as e:
            print(f"请求出现异常: {e}")
            return None

    def analyze_content(self):
        if self.current_page >= self.max_pages:
            print("已达到最大爬取页数，停止爬取。")
            return

        content = self.get_content()
        if content is None:
            return
        html = BeautifulSoup(content, 'lxml')
        for card in html.find_all('div', class_="card-wrap"):
            if 'mid' in str(card):
                uid = self.tool.extract_uid("https:" + card.find('div', class_="avator").find('a')['href'])
                age, location = self.get_user_detail_info(uid)
                mid = card['mid']
                detail_data = self.get_detail_data(mid)
                weibo_text = self.tool.clean_text(detail_data.get('text', '') if detail_data else "")
                self.tool.save_to_excel(FILENAMES, [age, location, weibo_text])

        next_page = html.find('a', class_="next")
        if next_page:
            next_url = 'https://s.weibo.com' + next_page['href']
            print(f'正在获取下一页: {next_url}')
            self.current_page += 1
            WeiboSpider(next_url, self.tool, self.driver, self.max_pages).analyze_content()
        else:
            print("没有更多页面可爬取。")

    def get_user_detail_info(self, uid):
        url = f"https://weibo.com/ajax/profile/detail?uid={uid}"
        try:
            data = requests.get(url, headers=self.tool.async_headers).json()
            if data.get('ok') in [1, '1']:
                user = data.get('data', {})
                birthday = user.get('birthday')
                age = ''
                if birthday:
                    try:
                        date_str = birthday.strip().split(' ')[0]
                        if len(date_str.split('-')) == 3:
                            birth_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
                            age = today.year - birth_date.year
                            if today.month < birth_date.month or (today.month == birth_date.month and today.day < birth_date.day):
                                age -= 1
                    except ValueError:
                        pass
                return age, user.get('ip_location', '').replace("IP属地：", "")
        except (requests.RequestException, json.JSONDecodeError):
            print(f"获取用户详细信息出错")
        return '', ''

    def get_detail_data(self, mid):
        url = f'https://weibo.com/ajax/detail?mid={mid}'
        try:
            data = requests.get(url, headers=self.tool.async_headers).json()
            return data['data'] if data.get('ok') in [1, '1'] else None
        except (requests.RequestException, json.JSONDecodeError, KeyError):
            print(f"获取微博详情出错")
            return None


if __name__ == '__main__':
    COOKIES = ''  # 这里可以留空，因为我们将手动登录获取 cookie
    from_time = '2024-07-24 00:00:00'
    to_time = '2024-07-24 6:00:00'
    # to_time = '2024-07-24 23:59:59'
    frequency = 6
    keyword = '巴黎奥运会'
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    FILENAMES = os.path.join(desktop_path, f'{"_".join(keyword)}2024.07.24-2024.08.11.xlsx')  # 进行数据存储

    tool = WeiboTool(COOKIES)
    tool.save_to_excel(FILENAMES, ['年龄', '所属IP地址', '微博内容'])

    # 配置 Chrome 浏览器驱动
    chrome_options = Options()
    # 如果你想在后台运行浏览器，可以取消注释下面这行
    # chrome_options.add_argument('--headless')
    service = Service('/Users/liujuncheng/PycharmProjects/start_api/改代码/chromedriver')
    driver = webdriver.Chrome(service=service, options=chrome_options)

    # 打开微博登录页面
    driver.get('https://weibo.com/login.php')
    print("请在打开的浏览器中手动登录微博，登录完成后按回车键继续...")
    input()

    # 获取登录后的 cookies
    cookies = driver.get_cookies()
    cookie_str = '; '.join([f'{cookie["name"]}={cookie["value"]}' for cookie in cookies])
    tool.weibo_content_headers['cookie'] = cookie_str
    tool.async_headers['cookie'] = cookie_str

    for start, end in tool.get_time_ranges(from_time, to_time, frequency):
        print(f'\n正在爬取时间段为: {start} ~ {end}')
        url = f'https://s.weibo.com/weibo?q={keyword}&typeall=1&suball=1&timescope=custom:{start}:{end}&Refer=g'
        print(url)  # 输出示例：custom:2024-07-24-0:2024-07-24-6
        WeiboSpider(url, tool, driver).analyze_content()
        sleep_time = random.randint(5, 10)
        print(f"本次请求完成，休眠 {sleep_time} 秒")
        time.sleep(sleep_time)

    # 关闭浏览器
    driver.quit()
